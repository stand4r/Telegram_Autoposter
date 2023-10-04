from telethon import errors
from telethon import TelegramClient, events
import datetime
from loguru import logger
import asyncio
from openpyxl import load_workbook


# ---------------------------------- TELEGRAM CLIENT INF ------------------------------------------------------------
api_id = 11283879
api_hash = '82c6793ac822d0ad61d08d3db2f62add'

# ------------------------------------ LOGGER -----------------------------------------------------------------------
logger.add("logs/logs.log", level='DEBUG')
logger.info("Время запуска скрипта = {}".format(datetime.datetime.now().strftime("Дата: %d/%m/%Y  Время: %H:%M:%S")))

# ---------------------------------- DICTS ---------------------------------------------------------------------------
bad_strings = open('bad.txt', 'r', encoding='utf8').readlines()

# ----------------------------------------- Временные -----------------------------------------------------------------
textForPosting = open('keyword.txt', 'r', encoding='utf8').readlines()
textForPosting_anime = open('keyword_anime.txt', 'r', encoding='utf8').readlines()


donors = []
mys = []
tags = []
wb = load_workbook('./source.xlsx')
sheet = wb["Лист1"]
count_channels = sheet.max_row
for x in range(2, count_channels):
    donors.append(sheet.cell(row=x, column=3).value)
    tags.append(sheet.cell(row=x, column=5).value)
    if sheet.cell(row=x, column=7).value is not None:
        mys.append(sheet.cell(row=x, column=7).value)
    else:
        mys.append(sheet.cell(row=x, column=6).value)


def log(level=None, myId=None, donorId=None):
    """
    :param donorId:
    :param myId:
    :param level:
        None -> start function
        0 -> success
        1 -> FloodWait Error
        2 -> Any Error
    :return: None
    """
    if level is None:
        logger.info("[*] Дата запуска =", datetime.datetime.now().strftime("Дата: %d/%m/%Y  Время: %H:%M:%S"))
        logger.info("[*] My channel = ", myId)
        logger.info("[*] Donor channel = ", donorId)
    elif level == 0:
        logger.info("[+] Сообщение успешно отправлено")
        logger.info("")
    elif level == 1:
        logger.debug("[!] FloodWait Error")
        logger.info("")
    elif level == 2:
        logger.debug("[!] Ошибка в работе скрипта")
        logger.info("")


with TelegramClient('Get', api_id, api_hash) as client:
    @client.on(events.NewMessage(chats=donors))
    async def message(event):
        index = donors.index(event.user_id)
        text = event.raw_text + '\n' + tags[index]
        if not [element for element in bad_strings if event.raw_text.lower().__contains__(element)]:
            log(id=mys[index], user=donors[index])
            try:
                if not event.grouped_id:
                    await client.send_message(
                        entity=mys[index],
                        file=event.message.media,
                        message=text,
                        parse_mode='md',
                        link_preview=False)
                    log(level=0)
                elif event.message.text and event.message.media is None and not event.grouped_id:
                    await client.send_message(
                        entity=mys[index],
                        message=text,
                        parse_mode='md',
                        link_preview=False)
                elif event.message.forward:
                    await event.message.forward_to(mys[index])
            except errors.FloodWaitError as e:
                log(level=1)
                await asyncio.sleep(e.seconds)
            except Exception:
                log(level=2)


    @client.on(events.Album(chats=donors))
    async def album(event):
        index = donors.index(event.user_id)
        text = event.original_update.message.message + '\n' + tags[index]
        if not [element for element in bad_strings if event.raw_text.lower().__contains__(element)]:
            log(id=mys[index], user=donors[index])
            try:
                await client.send_message(
                    entity=mys[index],
                    file=event.messages,
                    message=text,
                    parse_mode='md',
                    link_preview=False)
            except errors.FloodWaitError as e:
                log(level=1)
                await asyncio.sleep(e.seconds)
            except Exception:
                log(level=2)


    client.start()
    client.run_until_disconnected()
